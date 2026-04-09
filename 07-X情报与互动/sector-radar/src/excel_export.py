"""
Excel 导出模块
将 Pipeline 结果导出为格式化的 .xlsx 文件
"""

import logging
from datetime import datetime
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .sectors import Company, get_sector, SECTORS

logger = logging.getLogger(__name__)

# ── Styles ───────────────────────────────────
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
SECTOR_FILL = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
SECTOR_FONT = Font(name="Arial", size=11, bold=True, color="1B4F72")
NORMAL_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="2E86C1", underline="single")
SOURCE_COLORS = {
    "seed": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),       # 浅黄
    "bio_search": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),  # 浅蓝
    "following_network": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),  # 浅紫
}
THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
WRAP = Alignment(wrap_text=True, vertical="center")


def export_sector_excel(
    companies: List[Company],
    sector_id: str,
    filepath: str,
) -> str:
    """导出单个赛道的结果为 Excel"""
    sector = get_sector(sector_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{sector.name_cn}企业名单"

    # 列定义
    columns = ["#", "企业名称", "X 账号", "X 链接", "简介", "总部", "发现来源"]
    col_widths = [5, 25, 20, 35, 50, 12, 15]

    # 设置列宽
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{sector.name_cn} ({sector.name_en}) — 企业 X 账号名单"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="1B4F72")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 副标题：CARF 义务
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = f"CARF 报告义务：{sector.carf_obligation}"
    sub_cell.font = Font(name="Arial", size=10, italic=True, color="666666")
    sub_cell.alignment = Alignment(horizontal="center")

    # 统计行
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(columns))
    stats_cell = ws.cell(row=3, column=1)
    seed_count = sum(1 for c in companies if c.source == "seed")
    bio_count = sum(1 for c in companies if c.source == "bio_search")
    follow_count = sum(1 for c in companies if c.source == "following_network")
    stats_cell.value = (
        f"共 {len(companies)} 家 | "
        f"种子 {seed_count} | Bio搜索 {bio_count} | Following网络 {follow_count} | "
        f"生成时间 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    stats_cell.font = Font(name="Arial", size=9, color="888888")
    stats_cell.alignment = Alignment(horizontal="center")

    # 表头
    header_row = 5
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 25

    # 数据行
    for i, company in enumerate(companies, 1):
        row = header_row + i
        source_label = {
            "seed": "种子企业",
            "bio_search": "Bio搜索",
            "following_network": "Following网络",
        }.get(company.source, company.source)

        values = [
            i,
            company.name,
            f"@{company.x_handle}",
            f"https://x.com/{company.x_handle}",
            company.description,
            company.hq,
            source_label,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP

            # X 链接列设为超链接
            if col_idx == 4:
                cell.font = LINK_FONT
                cell.hyperlink = val

            # 来源列标色
            if col_idx == 7:
                fill = SOURCE_COLORS.get(company.source)
                if fill:
                    cell.fill = fill

    # 冻结窗格
    ws.freeze_panes = f"A{header_row + 1}"

    # 自动筛选
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(companies)}"

    # ── 第二个 Sheet：图例 ──
    ws2 = wb.create_sheet("图例说明")
    legend = [
        ("来源", "说明", "颜色"),
        ("种子企业", "内置的赛道头部企业（人工整理）", "浅黄"),
        ("Bio搜索", "通过 TwitterAPI.io search_user 的 Bio 关键词发现", "浅蓝"),
        ("Following网络", "从种子企业的 following 列表中发现的同赛道企业", "浅紫"),
    ]
    for r, (a, b, c) in enumerate(legend, 1):
        ws2.cell(row=r, column=1, value=a).font = NORMAL_FONT if r > 1 else HEADER_FONT
        ws2.cell(row=r, column=2, value=b).font = NORMAL_FONT if r > 1 else HEADER_FONT
        ws2.cell(row=r, column=3, value=c).font = NORMAL_FONT if r > 1 else HEADER_FONT
        if r > 1:
            source_key = ["seed", "bio_search", "following_network"][r - 2]
            ws2.cell(row=r, column=3).fill = SOURCE_COLORS[source_key]

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 10

    wb.save(filepath)
    logger.info(f"Excel exported to {filepath}")
    return filepath


def export_all_sectors_excel(
    all_results: Dict[str, List[Company]],
    filepath: str,
) -> str:
    """导出所有赛道到一个 Excel 文件（每个赛道一个 Sheet）"""
    wb = openpyxl.Workbook()

    # 概览 Sheet
    ws_overview = wb.active
    ws_overview.title = "概览"
    headers_ov = ["赛道", "英文名", "企业数", "种子", "Bio搜索", "Following网络", "CARF义务"]
    for col, h in enumerate(headers_ov, 1):
        cell = ws_overview.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, (sid, companies) in enumerate(all_results.items(), 2):
        sector = get_sector(sid)
        ws_overview.cell(row=i, column=1, value=sector.name_cn)
        ws_overview.cell(row=i, column=2, value=sector.name_en)
        ws_overview.cell(row=i, column=3, value=len(companies))
        ws_overview.cell(row=i, column=4, value=sum(1 for c in companies if c.source == "seed"))
        ws_overview.cell(row=i, column=5, value=sum(1 for c in companies if c.source == "bio_search"))
        ws_overview.cell(row=i, column=6, value=sum(1 for c in companies if c.source == "following_network"))
        ws_overview.cell(row=i, column=7, value=sector.carf_obligation)

    for col in range(1, 8):
        ws_overview.column_dimensions[get_column_letter(col)].width = [15, 30, 8, 8, 10, 14, 45][col - 1]

    # 每个赛道一个 Sheet
    columns = ["#", "企业名称", "X 账号", "X 链接", "简介", "总部", "发现来源"]
    col_widths = [5, 25, 20, 35, 50, 12, 15]

    for sector_id, companies in all_results.items():
        sector = get_sector(sector_id)
        ws = wb.create_sheet(title=sector.name_cn[:31])  # Sheet name max 31 chars

        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        for i, company in enumerate(companies, 1):
            source_label = {"seed": "种子企业", "bio_search": "Bio搜索", "following_network": "Following网络"}.get(company.source, company.source)
            row_data = [i, company.name, f"@{company.x_handle}", f"https://x.com/{company.x_handle}", company.description, company.hq, source_label]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=i + 1, column=col_idx, value=val)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                if col_idx == 4:
                    cell.font = LINK_FONT
                    cell.hyperlink = val
                if col_idx == 7:
                    fill = SOURCE_COLORS.get(company.source)
                    if fill:
                        cell.fill = fill

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(companies) + 1}"

    wb.save(filepath)
    logger.info(f"All-sectors Excel exported to {filepath}")
    return filepath
