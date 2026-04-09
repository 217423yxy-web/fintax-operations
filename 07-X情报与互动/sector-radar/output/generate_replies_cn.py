"""
生成带中文翻译回帖 + 知识库来源说明的新版 Excel
"""
import json
import time
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import translators as ts

# 从 generate_replies.py 导入回帖列表
sys.path.insert(0, "/Users/nightyoung/社媒运营工具/sector-radar/output")
from generate_replies import REPLIES

INPUT_FILE  = "/Users/nightyoung/社媒运营工具/sector-radar/output/tweets_with_kb.json"
OUTPUT_FILE = "/Users/nightyoung/社媒运营工具/sector-radar/output/tweets_with_replies_cn.xlsx"


def translate_reply(text: str) -> str:
    """将英文回帖翻译为中文，失败时返回空字符串"""
    try:
        result = ts.translate_text(text, translator='bing', to_language='zh')
        return result or ""
    except Exception as e:
        print(f"  [翻译失败] {e}")
        return ""


def format_kb_sources(kb_chunks: list) -> str:
    """将 kb_chunks 格式化为来源摘要"""
    if not kb_chunks:
        return ""
    lines = []
    for i, chunk in enumerate(kb_chunks[:3], start=1):
        score = chunk.get("score", 0)
        text  = chunk.get("text", "").strip()
        # 截取前 120 字符
        snippet = text[:120].replace("\n", " ")
        lines.append(f"[{i}] (相关度 {score:.2f}) {snippet}…")
    return "\n".join(lines)


def create_excel(tweets, replies, cn_replies, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "推文回帖建议"

    # 样式
    header_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    high_rel_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    mid_rel_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    row_fill_even = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    row_fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font     = Font(name="Arial", size=9)
    reply_en_font = Font(name="Arial", size=9, color="1F3864")
    reply_cn_font = Font(name="Arial", size=9, color="7B2C2C")
    src_font      = Font(name="Arial", size=8, color="666666", italic=True)
    data_align    = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align  = Alignment(horizontal="center", vertical="top")

    headers = [
        "账号", "推文（英文）", "推文（中文）",
        "回帖（英文）", "回帖（中文）",
        "字数", "相关度", "热度(likes)",
        "知识库来源摘要", "链接"
    ]
    col_widths = [18, 42, 30, 48, 35, 8, 10, 12, 55, 50]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    for row_idx, (tweet, reply_en, reply_cn) in enumerate(zip(tweets, replies, cn_replies), start=2):
        relevance = tweet.get("relevance", 0)

        if relevance >= 5:
            base_fill = high_rel_fill
        elif relevance >= 3:
            base_fill = mid_rel_fill
        elif row_idx % 2 == 0:
            base_fill = row_fill_even
        else:
            base_fill = row_fill_odd

        kb_source = format_kb_sources(tweet.get("kb_chunks", []))

        row_values = [
            tweet.get("account", ""),
            tweet.get("tweet_en", ""),
            tweet.get("tweet_cn", ""),
            reply_en,
            reply_cn,
            len(reply_en),
            relevance,
            tweet.get("likes", 0),
            kb_source,
            tweet.get("url", ""),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = base_fill
            cell.border = thin_border

            if col_idx == 4:   # 英文回帖
                cell.font      = reply_en_font
                cell.alignment = data_align
            elif col_idx == 5: # 中文回帖
                cell.font      = reply_cn_font
                cell.alignment = data_align
            elif col_idx == 9: # 知识库来源
                cell.font      = src_font
                cell.alignment = data_align
            elif col_idx in [6, 7, 8]:
                cell.font      = data_font
                cell.alignment = center_align
            else:
                cell.font      = data_font
                cell.alignment = data_align

        ws.row_dimensions[row_idx].height = 90

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(output_path)
    return len(tweets)


if __name__ == "__main__":
    with open(INPUT_FILE, encoding='utf-8') as f:
        tweets = json.load(f)

    total = len(tweets)
    print(f"共 {total} 条推文，{len(REPLIES)} 条回帖")

    # 翻译所有回帖
    cn_replies = []
    for i, reply in enumerate(REPLIES):
        print(f"[{i+1:02d}/{total}] 翻译中...", end=" ", flush=True)
        cn = translate_reply(reply)
        cn_replies.append(cn)
        print(cn[:60] if cn else "(失败)")
        time.sleep(0.5)  # 避免频率限制

    count = create_excel(tweets, REPLIES, cn_replies, OUTPUT_FILE)
    print(f"\n完成！已保存 {count} 行到:\n{OUTPUT_FILE}")
