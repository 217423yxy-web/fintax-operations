import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Pre-generated replies for all 71 tweets, indexed 0-70
# All replies are within 220 characters
REPLIES = [
    # 0 @Bybit_Official - yield from stablecoins/gold/Mantle
    "Yield-bearing stablecoins sit in a regulatory gray zone -- under MiCA and HK's stablecoin ordinance, returns from reserve assets trigger compliance obligations distinct from standard e-money rules.",
    # 1 @coinbase - borrow USDC to pay taxes instead of selling
    "Borrowing avoids disposal under IRS Notice 2014-21, but collateral must be tracked at FMV on pledge date -- inaccurate cost basis records are the single most common crypto audit trigger.",
    # 2 @coinbase - tokenization headlines
    "The SEC's position is consistent: wrapping a security in a token doesn't change its legal character. Reg D exemptions and reporting requirements still attach to the underlying instrument.",
    # 3 @BingXOfficial - Tether Big Four audit, BTC/Iran, BoM/CME
    "Tether's Big Four audit of $184B in reserves must address reserve composition and IFRS 9 / US GAAP fair value hierarchy treatment -- it's more complex than a standard attestation.",
    # 4 @bitget - CLARITY Act limits on stablecoin yields, CRCL drops
    "The CLARITY Act's yield restrictions target unregulated deposit substitutes -- similar to how MiCA caps stablecoin reserve returns to avoid destabilizing traditional banking. The US is converging on the same principle.",
    # 5 @MEXC_Official - flexible stablecoin savings, USDT/USDC yield
    "Stablecoin yield is taxable as ordinary income at receipt in the US, Singapore, and most major jurisdictions -- the flexible structure doesn't defer recognition, regardless of whether funds stay on-platform.",
    # 6 @BitMartExchange - SOL 7% / ETH 3.25% on-chain staking yields
    "Under IRS Rev. Rul. 2023-14, SOL and ETH staking rewards are ordinary income at FMV when received -- each reward creates a distinct taxable lot with its own cost basis, regardless of the flexible term.",
    # 7 @coinhako - compliance/security/infrastructure as exchange moat
    "Under CARF and DAC8, compliance infrastructure is the actual moat -- exchanges need reporting systems that serve both regulators and users meeting tax filing obligations across multiple jurisdictions.",
    # 8 @blockchain - 250+ tokenized stocks/ETFs in Blockchain Wallet
    "Tokenized stocks retain their securities classification regardless of the blockchain wrapper -- Reg D, Reg S, and Section 12(g) obligations remain intact, limiting 24/7 trading to non-US persons in most structures.",
    # 9 @Bybit_Official - stablecoin + gold earn carnival
    "Each asset class here triggers different tax treatment -- XAUT disposals are capital gains as a commodity, while stablecoin yield is ordinary income at receipt. The tax profile of a mixed portfolio matters.",
    # 10 @coinhako - ETF inflows, geopolitics, macro recap
    "Institutional ETF demand staying steady through volatility reflects one structural reality: funds under MiFID II or SEC frameworks have rigid rebalancing rules that mechanically dampen panic-selling behavior.",
    # 11 @binance - ETH soft staking, no lock-up
    "ETH staking rewards are ordinary income at receipt under IRS Rev. Rul. 2023-14 -- the no-lock-up structure doesn't defer that. Each epoch reward creates a separate taxable lot with its own cost basis.",
    # 12 @Gemini - NY tax hike on $1M+ earners
    "NY residents already face federal rates up to 37% on short-term crypto gains plus the 3.8% NIIT -- stacking a state hike on top makes NYC one of the highest-effective-rate crypto jurisdictions globally.",
    # 13 @LBank_Exchange - PRL/Perle Labs on-chain auditable AI data
    "On-chain data provenance addresses a real compliance gap -- under DAC8 and CARF, AI-generated financial analysis may soon face the same auditability scrutiny already applied to transaction data by reporting entities.",
    # 14 @Bitpanda - EU institutional on-chain infrastructure, MiCA/OP Stack
    "Building on MiCA's framework is the right foundation -- Article 35 requires CASPs to segregate client assets and hold capital buffers. An OP Stack environment needs those requirements mapped to smart contract controls.",
    # 15 @kucoincom - PRL/Perle Labs on-chain auditable AI infrastructure
    "On-chain auditable infrastructure is relevant to CARF -- CASPs must verify and report user transaction data, and machine-verifiable provenance reduces the reconciliation burden on reporting entities.",
    # 16 @MEXC_Official - 117 ONDO tokenized stocks on DCA/Convert
    "Each DCA purchase of tokenized stocks is a separate acquisition with its own cost basis -- in the UK (pool cost rules) or Ireland (FIFO), the annual reporting burden scales linearly with purchase frequency.",
    # 17 @Poloniex - 0 fees trading
    "Zero trading fees don't eliminate the tax -- every trade is a disposal event triggering capital gains, and under IRS/HMRC rules, fees adjust cost basis on each side of the trade even when those fees are zero.",
    # 18 @BitMartExchange - PRL/Perle Labs primary listing
    "On-chain auditable AI training data aligns with AML/CFT expectations under CARF and DAC8 -- tokens with verifiable data provenance have a stronger due diligence narrative than opaque token structures.",
    # 19 @WazirXIndia - SEC classifies 16 tokens as digital commodities, AI wallets
    "SEC's CLARITY classification of 16 tokens as digital commodities matters for tax reporting -- CFTC-regulated commodity gains carry different broker reporting thresholds than SEC-regulated securities.",
    # 20 @swissborg - HoF Days Geneva, DeFi/crypto derivatives
    "Under MiCA and ESMA guidance, crypto derivatives are classified as financial instruments -- meaning EMIR reporting and MiFID II transaction reporting apply as immediate obligations for EU-licensed platforms.",
    # 21 @bitfinex - BTC at $67K while equities slip, institutional bid
    "Under FASB ASU 2023-08, corporate Bitcoin holders must mark to fair value each period, creating direct EPS volatility -- which is exactly why institutional risk committees price in BTC's equity correlation so carefully.",
    # 22 @bitflyer - Japan crypto tax/ETF, Nikkei Money
    "Japan moving crypto gains from 55% 'miscellaneous income' to a 20% separate rate with loss carryforward would bring it in line with Singapore and Germany -- the most consequential crypto tax reform in Asia this decade.",
    # 23 @HashKeyExchange - Asia Connect Forum at Web3Festival
    "Under HK's VATP framework, licensed exchanges must maintain real-time order surveillance and report unusual patterns to the SFC -- that compliance floor differentiates regulated venues from offshore alternatives.",
    # 24 @MEXC_Official - tokenized stocks 24/7 DCA
    "Each automated stock DCA purchase creates a new tax lot -- in jurisdictions using FIFO (Ireland) or pool cost rules (UK), annual reporting burden scales linearly with purchase frequency and is often underestimated.",
    # 25 @Poloniex - trading fees are a 'tax on freedom'
    "Under IRS and HMRC guidance, fees on acquisition increase cost basis while fees on disposal reduce proceeds -- accurate fee tracking directly affects every capital gain calculation, even when fees are low.",
    # 26 @BitMartExchange - IZKY listing with rewards/airdrop
    "Deposit rewards and airdrops like IZKY's launch incentives are ordinary income at FMV on distribution date under IRS Rev. Rul. 2023-14 -- the 100% APY savings component creates additional recognition events throughout.",
    # 27 @WazirXIndia - weekly crypto recap: SEC, Solana stablecoin volume
    "Solana's $650B stablecoin volume falls within CARF's reporting scope -- above-threshold transactions trigger automatic exchange of financial information between jurisdictions, making scale a direct compliance multiplier.",
    # 28 @swissborg - SwissBorg MiCA compliance, #1 ranking
    "MiCA Article 70 requires documented conflict-of-interest policies for CASPs -- routing through 40+ venues also requires best execution documentation to satisfy MiCA's Article 27 client protection standards.",
    # 29 @okx - regulatory clarity building, DCA through volatility
    "CARF rolls out to 60+ jurisdictions by 2027, giving tax authorities automatic visibility into crypto holdings -- disciplined DCA record-keeping becomes essential as mismatches will be flagged automatically.",
    # 30 @HashKeyExchange - CC/Canton Network listing, professional investors
    "HK's VATP regime requires suitability assessments and enhanced due diligence for complex digital assets to professional investors -- similar to structured product rules, but with crypto-specific audit trails.",
    # 31 @Phemex_official - Astral Trading League PnL competition, USDT prize pool
    "USDT competition rewards are ordinary income when distributed -- platforms paying >$600 to US persons face 1099-MISC obligations, and non-US platforms reaching US participants may trigger FATCA withholding.",
    # 32 @okx - DAS2026 NYC, SEC token taxonomy, CFTC Innovation Task Force
    "CLARITY's decentralization classification test requires issuers to document governance and token distribution data now -- both the SEC and CFTC have signaled retroactive scrutiny of undocumented classification claims.",
    # 33 @swissborg - trigger orders live, stop-loss/take-profit automation
    "Each trigger order execution is a taxable disposal at the triggered price -- when stop-loss and take-profit both fire in a short window, wash sale-adjacent situations can arise in jurisdictions applying substance tests.",
    # 34 @Poloniex - Poloniex Ambassador announcement
    "Token-based ambassador compensation is ordinary income at FMV on receipt -- any subsequent appreciation when tokens are sold creates a separate capital gains event on top of that original income recognition.",
    # 35 @Official_Upbit - Korean-language EDGE/Definitive listing
    "Under Korea's Virtual Asset User Protection Act, exchanges listing new assets must segregate customer funds, carry insurance, and file SARs with the FIU -- each listing triggers a formal due diligence review.",
    # 36 @CoinDCX - crypto is legal as virtual digital asset in India
    "India's 30% flat rate under Section 115BBH plus 1% TDS per transaction under Section 194S makes it one of the highest-friction crypto tax regimes globally -- no loss offsets between VDA transactions.",
    # 37 @Bitso - ETH/USDC deposits via Base network
    "Mexico's SAT and Brazil's Receita Federal are moving toward CARF-aligned mandatory exchange reporting -- on-chain Base transactions will increasingly feed tax authority databases as those frameworks activate.",
    # 38 @HTX_Global - earn carnival, up to 20% APY on 28 cryptos
    "Promotional yield in crypto generates taxable income at FMV upon receipt -- HTX's +8% APY booster coupon creates recognition at each distribution event, and the coupon itself may be additional compensation income.",
    # 39 @swissborg - HoF Days Geneva (French tweet)
    "Under MiCA Title V, CASPs passporting across EU states must align product disclosure with local ESMA rules -- DeFi derivatives and crypto ETF forums at Geneva are increasingly compliance-driven.",
    # 40 @bitfinex - BTC $72K technical analysis, air gap to $82K
    "Under FASB ASU 2023-08, companies holding BTC must recognize fair value changes through earnings each period -- Bitcoin accumulation near $72K directly feeds into quarterly EPS in a way that didn't exist pre-2024.",
    # 41 @bitflyer - Japan separate capital gains taxation podcast (Japanese)
    "Japan's separation tax reform -- moving crypto from 55% 'miscellaneous income' to 20% with loss carryforward -- is the policy most likely to reverse capital outflows to Singapore among Japanese retail investors.",
    # 42 @HashKeyExchange - Web3Festival Asia Connect Forum
    "HK's VATP-licensed exchanges face monthly SFC surveillance reporting obligations -- HashKey's institutional access focus reflects demand for compliant on/off-ramp infrastructure that smaller offshore venues cannot match.",
    # 43 @Phemex_official - CEO market commentary
    "CARF's global rollout toward 2027 means exchanges face reporting obligations across 60+ jurisdictions simultaneously -- infrastructure for cross-border automatic information exchange needs to be built now.",
    # 44 @CoinoneOfficial - Korean XDC quiz event
    "Exchange quiz prizes in Korea are likely treated as miscellaneous income under the Income Tax Act -- with Korea's planned 20% VDA capital gains tax on amounts above KRW 2.5M, active users face compounding obligations.",
    # 45 @bitflyer - Japan crypto ETF discussion, FIN/SUM 2026 (Japanese)
    "Japan's FSA crypto ETF pathway through the Investment Trust Act amendment and tax reform must advance together -- ETF demand won't materialize while the rate remains at 55% 'miscellaneous income'.",
    # 46 @bitflyer - bitFlyer at FIN/SUM 2026, blockchain/stablecoin/DeFi/NFT panel
    "Japan's FSA must address stablecoins under the Payment Services Act, and DeFi, NFTs, DAOs under FIEA -- each has a distinct classification challenge the regulatory roadmap needs to resolve before 2030.",
    # 47 @bitflyer - Japan crypto separate taxation, 'miscellaneous income' reform
    "Japan's inability to offset crypto losses against other income has driven retail capital to Singapore -- moving to 20% flat rate with loss carryforward would structurally reverse that capital outflow almost immediately.",
    # 48 @MercadoBitcoin - stablecoin growth blog post (Portuguese)
    "Brazil's 90%+ stablecoin transaction share triggered the Central Bank's 100% reserve requirement and algorithmic stablecoin ban -- converging toward MiCA-level reserve and redemption standards for the LatAm market.",
    # 49 @coinbase - ETF perpetuals on SPY/QQQ for non-US traders
    "ETF perpetuals on equity indices outside the US may be treated as CFDs or derivatives -- with mark-to-market or realization treatment depending on jurisdiction, and equity-referenced gains potentially taxed differently.",
    # 50 @BingXOfficial - Nvidia AGI, Ethereum L1/L2, Australia pension crypto
    "Australia's $105B Hostplus pension exploring crypto is significant -- super funds are taxed at 15% on gains under ATO guidance, well below individual rates, making them potentially the most tax-efficient crypto vehicle.",
    # 51 @binance - margin modes explainer
    "Cross-margin liquidations can affect cost basis across all open positions simultaneously -- a tax tracking complexity that scales with account size and is significantly harder than isolated margin structures.",
    # 52 @LBank_Exchange - Web3 creator ownership, DeFi Tycoon spaces
    "Token-based creator revenue is ordinary income at distribution -- DAOs paying revenue to token holders create income at distribution date, not redemption, requiring real-time tax accrual tracking for compliance.",
    # 53 @LBank_Exchange - weekly risk-free trading, 100% loss protection
    "Risk-free promotions still create taxable events -- the initial trade is a disposal, and the platform's loss reimbursement is likely ordinary income, not a reversal. Users may owe tax even on net-zero campaigns.",
    # 54 @swissborg - SwissBorg #1 ranking, MiCA announcement
    "MiCA's CASP authorization requires ongoing compliance with capital ratios, Article 72 outsourcing controls, and Article 70 conflict-of-interest policies -- a continuous discipline, not a one-time approval.",
    # 55 @okx - DCA through regulatory noise
    "CARF across 60+ jurisdictions by 2027 means tax authorities receive standardized crypto data automatically -- compliant DCA record-keeping is necessary to avoid reconciliation gaps under automatic exchange.",
    # 56 @HashKeyExchange - CC Canton Network listing, professional investors
    "HK's VATP regime requires enhanced due diligence, suitability documentation, and quarterly reporting for complex digital assets to professional investors -- substantially more than standard VASP registration.",
    # 57 @Phemex_official - PnL competition, USDT pool, mystery boxes
    "USDT prize pools are ordinary income when received -- US platforms face 1099 reporting above $600, and globally, CRS and FATCA require withholding or reporting to participants' home jurisdiction tax authorities.",
    # 58 @okx - DAS2026, SEC taxonomy, CFTC task force
    "CLARITY's decentralization classification test means token issuers should document governance and distribution data proactively -- retroactive reclassification by either SEC or CFTC carries back-tax and penalty exposure.",
    # 59 @swissborg - trigger orders, stop-loss/take-profit automation
    "Automated trigger orders across 40+ venues each generate a taxable disposal -- without consolidated reporting, users face reconciliation gaps when computing annual capital gains across multi-venue execution.",
    # 60 @Poloniex - CE_YIYI ambassador announcement
    "Ambassador token compensation is ordinary income at FMV on receipt -- if the token subsequently appreciates, the sale triggers a separate capital gains event layered on top of the original income recognition.",
    # 61 @Official_Upbit - EDGE listing, KRW market (Korean)
    "Korea's Virtual Asset User Protection Act requires asset segregation and insurance for listed tokens -- with the planned 20% VDA capital gains tax above KRW 2.5M, tax reporting infrastructure is needed now.",
    # 62 @CoinDCX - 'crypto is legal as VDA' ad
    "India's VDA status is clear, but the 30% flat rate under Section 115BBH with no deductions for losses from other VDA transactions makes India structurally punitive for active traders at any volume level.",
    # 63 @Bitso - ETH/USDC deposits via Base, LatAm focus
    "Bridging ETH from mainnet to Base may constitute a taxable disposal in many jurisdictions -- treating the bridge as a sale and reacquisition, with distinct cost basis on each side, is the conservative compliance view.",
    # 64 @HTX_Global - LIT flexible yield, 12% APY
    "Altcoin yield at 12% APY is ordinary income at FMV on each distribution -- not when you sell the earned tokens. Users can face tax liability on yield even if the token declines before disposal.",
    # 65 @MEXC_Official - MEXC at Mexico Blockchain Week
    "Mexico's SAT is expanding CARF-aligned reporting for FSTOs under the Fintech Law -- exchange compliance infrastructure is becoming a front-line regulatory issue, not just back-office, as mandatory reporting scales up.",
    # 66 @Bybit_Official - new user 666% APR promo
    "Promotional USDT yield is ordinary income at receipt regardless of APR -- extremely high short-term rates create phantom income risk where tax liability can exceed actual withdrawable cash in the promotion window.",
    # 67 @bitflyer - Japan separate capital gains taxation podcast (Japanese)
    "Japan's 55% maximum on crypto 'miscellaneous income' with no cross-asset loss offset drives capital to Singapore -- a 20% separate rate with loss carryforward would structurally reverse that outflow.",
    # 68 @coincheckjp - Digital Asset Forum 2026, crypto ETF/on-chain TradFi
    "Japan's Investment Trust Act amendment would classify crypto spot ETFs as specified asset class funds -- requiring the same disclosure standards as conventional trusts, well beyond current VASP-level obligations.",
    # 69 @Bitstamp - January lending report, 11 borrowers, collateral/D-E ratios
    "Bitstamp's lending transparency is exactly what CARF envisages for crypto lenders -- both interest income earned by lenders and collateral liquidation events must be reported to tax authorities in the lender's residence.",
    # 70 @WhiteBit - ZEN Pay for PLN transactions, 1% fee
    "WhiteBit's PLN on/off-ramp triggers AML obligations under Poland's AMLD5 -- VASPs must report suspicious transactions above PLN 15,000 to GIIF, and the 1% fee adjusts cost basis on both sides of each transaction.",
]


def load_tweets(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def create_excel(tweets, replies, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "推文回帖建议"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    headers = ["账号", "推文（英文）", "推文（中文）", "回帖建议", "字数", "相关度", "热度(likes)", "链接"]
    col_widths = [18, 48, 35, 52, 8, 10, 12, 52]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    high_rel_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    mid_rel_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    row_fill_even = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    row_fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data_font   = Font(name="Arial", size=9)
    reply_font  = Font(name="Arial", size=9, color="1F3864")
    data_align  = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    for row_idx, (tweet, reply) in enumerate(zip(tweets, replies), start=2):
        relevance = tweet.get("relevance", 0)
        likes = tweet.get("likes", 0)

        if relevance >= 5:
            base_fill = high_rel_fill
        elif relevance >= 3:
            base_fill = mid_rel_fill
        elif row_idx % 2 == 0:
            base_fill = row_fill_even
        else:
            base_fill = row_fill_odd

        char_count = len(reply)

        row_values = [
            tweet.get("account", ""),
            tweet.get("tweet_en", ""),
            tweet.get("tweet_cn", ""),
            reply,
            char_count,
            relevance,
            likes,
            tweet.get("url", ""),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = base_fill
            cell.border = thin_border

            if col_idx == 4:
                cell.font = reply_font
                cell.alignment = data_align
            elif col_idx in [5, 6, 7]:
                cell.font = data_font
                cell.alignment = center_align
            else:
                cell.font = data_font
                cell.alignment = data_align

        ws.row_dimensions[row_idx].height = 80

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(output_path)
    return len(tweets)


if __name__ == "__main__":
    input_path  = "/Users/nightyoung/社媒运营工具/sector-radar/output/tweets_with_kb.json"
    output_path = "/Users/nightyoung/社媒运营工具/sector-radar/output/tweets_with_replies.xlsx"

    tweets = load_tweets(input_path)
    print(f"Loaded {len(tweets)} tweets")
    print(f"Prepared {len(REPLIES)} replies")

    if len(tweets) != len(REPLIES):
        print(f"WARNING: tweet count ({len(tweets)}) != reply count ({len(REPLIES)})")

    over_limit = [(i, len(r)) for i, r in enumerate(REPLIES) if len(r) > 220]
    if over_limit:
        print("Replies OVER 220 chars:")
        for idx, c in over_limit:
            print(f"  [{idx}] {c} chars: {REPLIES[idx]}")
    else:
        max_len = max(len(r) for r in REPLIES)
        print(f"All replies within 220 chars. Max length: {max_len}")

    count = create_excel(tweets, REPLIES, output_path)
    print(f"Done. Saved {count} rows to {output_path}")
