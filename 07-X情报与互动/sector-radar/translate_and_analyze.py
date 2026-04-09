#!/usr/bin/env python3
"""
读取 CEX 推文，用 Bing 翻译到中文（Microsoft 翻译，非 Google，免费无需 API Key），
标注 FinTax 互动机会，生成 Excel。
"""
import json, time
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = Path(__file__).parent / "output/cex_tweets_twikit_20260326_1335.json"
OUTPUT_FILE = Path(__file__).parent / "output/cex_tweets_cn_analyzed.xlsx"

# ── FinTax 互动机会判断 ─────────────────────────────────────────────────────────
ENGAGE_KEYWORDS = [
    "tax", "taxes", "taxable", "irs", "carf", "1099", "dac8", "vat", "fiscal",
    "compliance", "regulation", "regulatory", "sec ", "cftc", "clarity act",
    "legal", "license", "licensed", "regulator", "audit",
    "accounting", "financial report", "balance sheet", "proof of reserve",
    "lending report", "debt-to-equity", "portfolio",
    "institutional", "tokenization", "tokenized", "tradfi",
    "staking", "staking reward", "yield", "earn reward",
    "equity perp", "stock perp", "etf perpetual",
]

# 高优先级推文的人工标注互动角度（key 为推文 URL 中的 tweet ID 片段）
ENGAGE_ANGLES = {
    # ⭐⭐⭐ 最高优先级 — 直接触及税务/合规核心议题
    "2036816814655828190": (
        "【直接税务话题 ⭐⭐⭐】这条推文讨论了持有加密货币时的税务循环问题（卖币缴税困境）。"
        "FinTax 可以回复：「除了借贷方案，企业还可以用 FinTax 提前规划纳税时点——我们的 "
        "CARF/1099DA 申报平台帮交易所和机构提前识别应税事件，避免被动触发纳税义务。」"
    ),
    "2036761012238778539": (
        "【代币化资产会计 ⭐⭐⭐】讨论机构资产代币化。"
        "FinTax 可以回复：「代币化资产带来的不只是价格发现，还有会计分类难题——究竟是证券、商品还是无形资产？"
        "FinTax Suite 已支持 IFRS/GAAP 框架下的代币化资产核算。」"
    ),
    "2036551866545295603": (
        "【SEC/CFTC 监管清晰化 ⭐⭐⭐】OKX 提到 SEC 代币分类清晰化 + CFTC 创新任务组。"
        "FinTax 可以回复：「监管清晰度提升意味着更严格的申报义务——CARF、DAC8 已在多个市场强制落地。"
        "FinTax 正在帮助交易所做好这件事。」"
    ),
    "2036697918787510668": (
        "【稳定币监管/CLARITY Act ⭐⭐⭐】CLARITY Act 限制稳定币收益，直接触发合规/会计议题。"
        "FinTax 可以回复：「稳定币收益的监管趋严，意味着交易所需要更精准的税务核算体系——"
        "FinTax 支持稳定币业务的完整会计处理和合规申报。」"
    ),
    "2025945289467515186": (
        "【财务透明度/借贷报告 ⭐⭐⭐】Bitstamp 发布借贷月报，展示财务透明度。"
        "FinTax 可以回复：「这正是机构级财务透明度的实践——FinTax Suite 帮助交易所和借贷机构"
        "自动化生成符合审计标准的投资组合报告，降低手工整理成本 80%+。」"
    ),
    # ⭐⭐ 高优先级 — 间接相关，可以从 FinTax 角度切入
    "2036930119272943860": (
        "【HashKey 客户关系 ⭐⭐】HashKey 是 FinTax 客户，可转推或点赞表示支持，"
        "评论时可提及双方的合作关系以增加可信度。"
    ),
    "2036261988767662082": (
        "【HashKey Web3Festival 监管论坛 ⭐⭐】涉及监管议题的论坛。"
        "FinTax 可以合规专家身份参与讨论，介绍 CARF/DAC8 落地现状。"
    ),
    "2036805914452234528": (
        "【ETH 软质押税务 ⭐⭐】Binance ETH 软质押。"
        "FinTax 可以评论：「质押收益在多数司法管辖区构成应税所得——"
        "FinTax 帮助持有 ETH 的机构准确追踪每笔质押奖励的税务成本基础。」"
    ),
    "2036745581305921912": (
        "【TradFi 股权永续合约 ⭐⭐】Binance 推出 Meta/NVDA/Google TradFi 永续合约。"
        "FinTax 可以评论：「TradFi 股票合约在加密平台上交易，带来全新的会计分类挑战——"
        "FinTax 支持混合型资产的跨境税务处理。」"
    ),
    "2037009557214621900": (
        "【代币化股票 ⭐⭐】MEXC 推出 ONDO 代币化股票（117 支）。"
        "FinTax 可以评论：「117 支代币化股票上线，意味着企业会计需要同时处理两套资产逻辑——"
        "FinTax 已在帮客户处理代币化证券的 IFRS 合规核算。」"
    ),
    "2036856285157576782": (
        "【ETF 永续合约 ⭐⭐】Coinbase ETF 永续合约上线。"
        "类似 TradFi 混合资产话题，FinTax 可从会计分类角度切入。"
    ),
    "2018933562951827733": (
        "【交易所财务报告透明度 ⭐⭐】Bithumb 公开财务报告，展示透明度。"
        "FinTax 可以评论：「财务透明度是机构信任的基础——FinTax Suite 帮助交易所"
        "自动化生成符合审计标准的财务报告，降低合规成本。」"
    ),
    "2036427471537459305": (
        "【250+ 代币化资产 ⭐⭐】Blockchain.com 新增 250+ 代币化股票/ETF。"
        "FinTax 可以评论：「大规模代币化资产上线，交易所面临前所未有的会计复杂度——"
        "FinTax 已支持代币化证券在 IFRS/GAAP 框架下的自动化核算。」"
    ),
    # ⭐ 标准优先级 — 相关行业动态
    "2036471866545295603": (
        "【OKX × NYSE 市场结构合作 ⭐】"
        "FinTax 可从合规基础设施角度切入，介绍交易所与 TradFi 合作带来的申报合规新挑战。"
    ),
}


def contains_keyword(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in ENGAGE_KEYWORDS)


def get_engage_angle(url: str, text: str) -> tuple[bool, str]:
    """返回 (是否可互动, 互动角度建议)"""
    for frag, angle in ENGAGE_ANGLES.items():
        if frag in url:
            return True, angle

    if not contains_keyword(text):
        return False, ""

    t = text.lower()
    if any(k in t for k in ["tax", "taxes", "taxable", "1099", "carf", "dac8"]):
        return True, "【税务相关 ⭐】FinTax 可以「加密税务合规专家」身份评论，分享 CARF/1099DA/DAC8 申报平台的实际价值。"
    if any(k in t for k in ["staking", "staking reward"]):
        return True, "【质押奖励税务 ⭐】质押奖励在多数司法管辖区为应税所得，FinTax Suite 自动识别每笔质押事件的税务成本基础。"
    if any(k in t for k in ["compliance", "regulation", "regulatory", "sec ", "cftc"]):
        return True, "【监管/合规 ⭐】FinTax 可以「加密合规基础设施」角度切入，介绍 CARF/DAC8 落地情况。"
    if any(k in t for k in ["tokenization", "tokenized"]):
        return True, "【代币化资产会计 ⭐】FinTax 可以评论代币化资产在 IFRS/GAAP 框架下的处理逻辑。"
    if any(k in t for k in ["audit", "proof of reserve", "lending report", "financial report"]):
        return True, "【财务透明度/审计 ⭐】FinTax 帮助交易所生成符合审计标准的财务报告，SOC 1/2 Type II 认证。"
    if any(k in t for k in ["institutional", "tradfi"]):
        return True, "【机构/TradFi ⭐】FinTax 可评论机构采用加密资产后面临的会计/税务挑战及解决方案。"
    return True, "【相关合规议题 ⭐】FinTax 可以行业专家身份发表专业评论。"


# ── 翻译（Microsoft Bing，非 Google） ──────────────────────────────────────────
try:
    import translators as _ts
    _TS_AVAILABLE = True
except ImportError:
    _TS_AVAILABLE = False


def is_english(text: str) -> bool:
    """粗略判断推文是否主要为英文（可翻译）"""
    if not text:
        return False
    ascii_chars = sum(1 for c in text if ord(c) < 128)
    return ascii_chars / len(text) > 0.85


def translate_text(text: str, max_len=4000) -> str:
    """用 Bing 翻译英文推文到中文"""
    if not text or not text.strip():
        return ""
    if not is_english(text):
        return "[非英文，保留原文]"
    if not _TS_AVAILABLE:
        return "[未安装 translators 库]"
    if len(text) > max_len:
        text = text[:max_len] + "..."
    try:
        result = _ts.translate_text(text, translator="bing", to_language="zh")
        return result
    except Exception as e:
        return f"[翻译失败: {str(e)[:60]}]"


# ── Excel 样式 ──────────────────────────────────────────────────────────────────
def make_styles():
    header_fill  = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
    engage_fill  = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    alt_fill     = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal_font  = Font(name="Arial", size=10)
    engage_font  = Font(name="Arial", size=10, bold=True)
    thin_border  = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    return header_fill, engage_fill, alt_fill, header_font, normal_font, engage_font, thin_border


def write_header(ws, headers, col_widths, header_fill, header_font, thin_border):
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


# ── 主程序 ──────────────────────────────────────────────────────────────────────
def main():
    with open(INPUT_FILE, encoding="utf-8") as f:
        tweets = json.load(f)
    print(f"共 {len(tweets)} 条推文，使用 Bing 翻译（Microsoft，非 Google）...\n", flush=True)

    header_fill, engage_fill, alt_fill, header_font, normal_font, engage_font, thin_border = make_styles()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全部推文（含翻译）"

    headers1    = ["账号", "发布时间", "原文（英文）", "中文翻译", "可互动?", "互动角度建议", "点赞", "转推", "评论", "浏览量", "链接"]
    col_widths1 = [15,      18,          52,              52,          8,         48,              8,     8,     8,     10,    52]
    write_header(ws, headers1, col_widths1, header_fill, header_font, thin_border)

    engage_tweets = []
    row_idx = 2

    for i, tweet in enumerate(tweets):
        if i % 30 == 0:
            print(f"  处理 {i+1}/{len(tweets)}...", flush=True)

        text = tweet.get("text", "")
        chinese_text = translate_text(text)
        if is_english(text):
            time.sleep(0.5)  # 避免触发 Bing 频率限制

        engage, angle = get_engage_angle(tweet["url"], text)
        if engage:
            engage_tweets.append({**tweet, "chinese_text": chinese_text, "engage_angle": angle})

        row_data = [
            f"@{tweet['handle']}",
            tweet["date"][:16],
            text,
            chinese_text,
            "✅ 是" if engage else "",
            angle,
            tweet["likes"],
            tweet["retweets"],
            tweet["replies"],
            int(tweet["views"]) if str(tweet["views"]).isdigit() else tweet["views"],
            tweet["url"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = engage_font if engage else normal_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if engage:
                cell.fill = engage_fill
            elif row_idx % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 60
        row_idx += 1

    # ── Sheet 2: 互动机会汇总 ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("⭐ 互动机会汇总")
    headers2    = ["账号", "发布时间", "中文翻译", "互动角度建议", "点赞", "浏览量", "链接"]
    col_widths2 = [15,      18,          58,              58,          8,     10,    55]
    write_header(ws2, headers2, col_widths2, header_fill, header_font, thin_border)

    engage_tweets.sort(key=lambda x: int(x["likes"]), reverse=True)

    for i, t in enumerate(engage_tweets, 2):
        row_data = [
            f"@{t['handle']}",
            t["date"][:16],
            t["chinese_text"],
            t["engage_angle"],
            t["likes"],
            int(t["views"]) if str(t["views"]).isdigit() else t["views"],
            t["url"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col_idx, value=value)
            cell.font = normal_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = alt_fill
        ws2.row_dimensions[i].height = 80

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 完成！{len(tweets)} 条推文，{len(engage_tweets)} 条可互动", flush=True)
    print(f"输出: {OUTPUT_FILE}", flush=True)


if __name__ == "__main__":
    main()
